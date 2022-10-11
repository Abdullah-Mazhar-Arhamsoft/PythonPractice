import openpyxl

wb = openpyxl.load_workbook("transaction.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["a1"]
# print(cell.value)
# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)
column = sheet["a"]
cells = sheet["a:c"]
print(cells)
sheet["a1:c3"]
sheet[1:3]

sheet.append([1, 2, 3])
wb.save("transaction2.xlsx")
